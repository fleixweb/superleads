#!/usr/bin/env python3
"""Run Code Slice AG evals for single-customer background research."""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
CASES = ROOT / "evals" / "cases" / "customer_background_research_cases.json"
VALIDATOR = ROOT / "scripts" / "validate_research_graph.py"
WORKBOOK = ROOT / "scripts" / "export_workbook.py"
MARKDOWN = ROOT / "scripts" / "export_superleads_markdown.py"
USER_VISIBLE = ROOT / "scripts" / "validate_superleads_user_visible_output.py"


def _run(cmd: list[str], expect: int) -> dict[str, Any]:
    env = dict(os.environ)
    env["PYTHONDONTWRITEBYTECODE"] = "1"
    proc = subprocess.run(cmd, cwd=str(ROOT), text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, env=env)
    return {"cmd": cmd, "returncode": proc.returncode, "expected": expect, "ok": proc.returncode == expect, "output": proc.stdout}


def _load_cases() -> list[dict[str, Any]]:
    payload = json.loads(CASES.read_text(encoding="utf-8"))
    return [case for case in payload.get("cases", []) if isinstance(case, dict)]


def _assert_contains(text: str, needles: list[Any]) -> list[str]:
    return [str(needle) for needle in needles if str(needle) not in text]


def _assert_absent(text: str, needles: list[Any]) -> list[str]:
    return [str(needle) for needle in needles if str(needle) in text]


def _codes_missing(output: str, codes: list[Any]) -> list[str]:
    return [str(code) for code in codes if str(code) not in output]


def _validate_fixture(py: str, fixture: Path, case: dict[str, Any]) -> dict[str, Any]:
    expect = 0 if case.get("validate", case.get("expected", "pass")) == "pass" else 1
    result = _run([py, str(VALIDATOR), str(fixture), "--format", "json"], expect)
    missing = _codes_missing(str(result.get("output", "")), list(case.get("expected_error_codes", []))) if expect != 0 else []
    if missing:
        result["ok"] = False
        result["returncode"] = 1
        result["output"] += "\nmissing expected error codes: " + ", ".join(missing)
    return result


def _export_background(py: str, fixture: Path, tmp_path: Path, index: int, case: dict[str, Any]) -> dict[str, Any]:
    expect = 0 if case.get("export_background", "pass") == "pass" else 1
    out_dir = tmp_path / f"background_export_{index}"
    result = _run([py, str(WORKBOOK), str(fixture), "--output-dir", str(out_dir), "--mode", "background", "--format", "csv"], expect)
    if expect != 0 or not result.get("ok"):
        return result
    required_sheets = {
        "客户一眼看懂.csv",
        "客户、品牌与关联方.csv",
        "我们看到的业务机会.csv",
        "怎么联系、先找谁.csv",
        "跟进前要注意什么.csv",
        "信息从哪里来.csv",
    }
    required_sheets.update(str(item) for item in case.get("conditional_csv_sheets", []) if isinstance(item, str))
    produced = {path.name for path in out_dir.glob("*.csv")}
    problems: list[str] = []
    if produced != required_sheets:
        problems.append(f"unexpected CSV sheets: {sorted(produced)}")
    haystack = "\n".join(
        path.name + "\n" + path.read_text(encoding="utf-8-sig")
        for path in sorted(out_dir.glob("*.csv"))
    )
    missing = _assert_contains(haystack, list(case.get("csv_must_contain", [])))
    hits = _assert_absent(haystack, list(case.get("must_not_contain", [])))
    if missing or hits:
        problems.append(f"CSV assertions failed missing={missing} forbidden_hits={hits}")
    if problems:
        result["ok"] = False
        result["returncode"] = 1
        result["output"] += "\n" + "; ".join(problems)
        return result

    xlsx_needles = [str(item) for item in case.get("xlsx_must_contain", []) if isinstance(item, str)]
    if xlsx_needles:
        xlsx_dir = tmp_path / f"background_xlsx_{index}"
        xlsx_result = _run([py, str(WORKBOOK), str(fixture), "--output-dir", str(xlsx_dir), "--mode", "background", "--format", "xlsx"], 0)
        xlsx_files = list(xlsx_dir.glob("*.xlsx"))
        xlsx_text = ""
        if xlsx_files:
            try:
                from openpyxl import load_workbook  # type: ignore
                workbook = load_workbook(xlsx_files[0], read_only=True, data_only=True)
                xlsx_text = "\n".join(
                    f"{sheet}\n" + "\n".join(" | ".join(str(cell or "") for cell in row) for row in workbook[sheet].iter_rows(values_only=True))
                    for sheet in workbook.sheetnames
                )
            except Exception as exc:
                xlsx_result["ok"] = False
                xlsx_result["output"] += f"\nXLSX read failed: {exc}"
        missing_xlsx = _assert_contains(xlsx_text, xlsx_needles)
        if missing_xlsx or not xlsx_result.get("ok"):
            result["ok"] = False
            result["returncode"] = 1
            result["output"] += f"\nXLSX assertions failed missing={missing_xlsx}\n{xlsx_result.get('output', '')}"
    return result


def _markdown_delivery(py: str, fixture: Path, tmp_path: Path, index: int, case: dict[str, Any]) -> dict[str, Any]:
    expect = 0 if case.get("markdown_delivery", "pass") == "pass" else 1
    out = tmp_path / f"background_{index}.md"
    result = _run([py, str(MARKDOWN), str(fixture), "--route", "customer_background_research", "--output", str(out), "--format", "json"], expect)
    if expect != 0 or not result.get("ok"):
        return result
    if not out.exists():
        result["ok"] = False
        result["returncode"] = 1
        result["output"] += "\nMarkdown output was not written"
        return result
    visible = _run([py, str(USER_VISIBLE), str(out), "--route", "customer_background_research", "--min-tables", "6", "--format", "json"], 0)
    text = out.read_text(encoding="utf-8")
    missing = _assert_contains(text, list(case.get("must_contain", [])))
    hits = _assert_absent(text, list(case.get("must_not_contain", [])))
    if missing or hits or not visible.get("ok"):
        result["ok"] = False
        result["returncode"] = 1
        result["output"] += "\n" + str(visible.get("output", ""))
        result["output"] += f"\nMarkdown assertions failed missing={missing} forbidden_hits={hits}"
    return result


def _case(py: str, case: dict[str, Any], tmp_path: Path, index: int) -> dict[str, Any]:
    fixture = ROOT / "evals" / "fixtures" / str(case.get("fixture", ""))
    validate = _validate_fixture(py, fixture, case)
    export = {"ok": True, "output": "skipped"}
    markdown = {"ok": True, "output": "skipped"}
    if validate.get("ok") and case.get("expected", "pass") == "pass":
        export = _export_background(py, fixture, tmp_path, index, case)
        markdown = _markdown_delivery(py, fixture, tmp_path, index, case)
    return {
        "name": str(case.get("name", fixture.name)),
        "fixture": str(fixture.relative_to(ROOT)),
        "validate": validate,
        "export": export,
        "markdown": markdown,
        "ok": bool(validate.get("ok")) and bool(export.get("ok")) and bool(markdown.get("ok")),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--suite", choices=["all"], default="all")
    args = parser.parse_args()
    py = sys.executable
    results: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        for index, case in enumerate(_load_cases(), start=1):
            results.append(_case(py, case, tmp_path, index))
    total = len(results)
    passed = sum(1 for result in results if result.get("ok"))
    summary = {"suite": args.suite, "total": total, "passed": passed, "failed": total - passed, "results": results}
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if passed == total else 1


if __name__ == "__main__":
    raise SystemExit(main())
