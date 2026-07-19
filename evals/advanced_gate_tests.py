#!/usr/bin/env python3
"""Regression tests for formal-delivery bypasses found in independent reviews."""
from __future__ import annotations

import argparse
import copy
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Callable


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
BASE_GRAPH = ROOT / "evals" / "fixtures" / "pass_minimal_graph.json"
USER_FILE_PDF = ROOT / "evals" / "fixtures" / "pass_user_provided_pdf_claim.json"
USER_FILE_SHEET = ROOT / "evals" / "fixtures" / "pass_user_provided_spreadsheet_contact.json"
DATASET_CONTACT = ROOT / "evals" / "fixtures" / "pass_user_business_dataset_contact_with_note.json"
CORRESPONDENCE_CONTACT = ROOT / "evals" / "fixtures" / "pass_correspondence_export_contact_with_note.json"
VISUAL_CANDIDATE = ROOT / "evals" / "fixtures" / "pass_visual_reference_candidate_only.json"
CONNECTED_INQUIRY = ROOT / "evals" / "fixtures" / "pass_connected_inbound_inquiry.json"
MAIL_ADAPTER_INPUT = ROOT / "evals" / "fixtures" / "mail_read_normalized_input.json"
sys.path.insert(0, str(SCRIPTS))
from _superleads_common import graph_hash, is_safe_public_http_url, review_subject_hash  # noqa: E402
from export_workbook import build_initial_sheets, build_sheets  # noqa: E402
sys.path.insert(0, str(ROOT / "evals"))
from run_evals import _load_fixture_graph  # noqa: E402


def _run(command: list[str]) -> subprocess.CompletedProcess[str]:
    environment = dict(os.environ)
    environment["PYTHONDONTWRITEBYTECODE"] = "1"
    return subprocess.run(command, cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, env=environment)


def _write_graph(path: Path, graph: dict[str, Any]) -> None:
    path.write_text(json.dumps(graph, ensure_ascii=False, indent=2), encoding="utf-8")


def _formal_results(graph: dict[str, Any], directory: Path, name: str) -> tuple[subprocess.CompletedProcess[str], subprocess.CompletedProcess[str], subprocess.CompletedProcess[str]]:
    graph_path = directory / f"{name}.json"
    export_dir = directory / name
    _write_graph(graph_path, graph)
    validate = _run([sys.executable, "-B", str(SCRIPTS / "validate_research_graph.py"), str(graph_path), "--format", "json"])
    audit = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(graph_path), "--delivery-status", "standard_development_list", "--format", "json"])
    export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(export_dir), "--mode", "standard", "--format", "csv", "--manifest", str(export_dir / "manifest.json")])
    return validate, audit, export


def _must_block(name: str, graph: dict[str, Any], directory: Path) -> list[str]:
    validate, audit, export = _formal_results(graph, directory, name)
    errors = []
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0:
            errors.append(f"{name}: {label} unexpectedly passed\n{result.stdout}")
    return errors


def _base() -> dict[str, Any]:
    return json.loads(BASE_GRAPH.read_text(encoding="utf-8"))


def _append_unassigned(graph: dict[str, Any]) -> None:
    graph["unassigned_contact_leads"].append({
        "unassigned_contact_lead_id": "unassigned_001",
        "contact_id": "contact_001",
        "reason": "requires manual sourcing",
        "suggested_manual_check": "Find a public source.",
    })


def _refresh_current_attestation(graph: dict[str, Any]) -> None:
    """Simulate a current independent review after a test-only data change."""
    attestations = graph.get("review_attestations")
    if not isinstance(attestations, list) or not attestations or not isinstance(attestations[0], dict):
        return
    subject_hash = review_subject_hash(graph)
    attestations[0]["input_graph_hash"] = subject_hash
    attestations[0]["reviewed_subject_hash"] = subject_hash


def _assert_self_review_disclosure(directory: Path) -> list[str]:
    graph = _base()
    graph["runs"][0]["review_mode"] = "self_review_fallback"
    validate, audit, export = _formal_results(graph, directory, "self_review_disclosure")
    errors = []
    if any(result.returncode != 0 for result in (validate, audit, export)):
        return [f"self_review_disclosure: expected formal export to pass\n{audit.stdout}\n{export.stdout}"]
    try:
        if not json.loads(audit.stdout).get("disclosure_required"):
            errors.append("self_review_disclosure: audit must require disclosure")
    except json.JSONDecodeError:
        errors.append("self_review_disclosure: audit did not return JSON")
    disclosure_sheet = (directory / "self_review_disclosure" / "风险与说明.csv").read_text(encoding="utf-8-sig")
    if "本次未运行独立复核，建议在使用前进行人工确认。" not in disclosure_sheet:
        errors.append("self_review_disclosure: workbook risk sheet lacks self_review_fallback disclosure")
    return errors


def _assert_phase2_provenance_and_searchlog(directory: Path) -> list[str]:
    """Directly pressure the new provenance and candidate-only search gates."""
    errors: list[str] = []
    geography = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
    validate, audit, export = _formal_results(geography, directory, "phase2_geography_control")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        errors.append(f"phase2_geography_control: expected validate/audit/export pass\n{validate.stdout}\n{audit.stdout}\n{export.stdout}")
    else:
        try:
            if not json.loads(audit.stdout).get("disclosure_required"):
                errors.append("phase2_geography_control: declared review must require audit disclosure")
        except json.JSONDecodeError:
            errors.append("phase2_geography_control: audit did not return JSON")
        payload = (directory / "phase2_geography_control" / "manifest.json").read_text(encoding="utf-8")
        csv_text = "\n".join(path.read_text(encoding="utf-8-sig") for path in (directory / "phase2_geography_control").glob("*.csv"))
        disclosure = "本次复核由独立会话声明完成，未获得平台身份验证。"
        if disclosure not in payload or disclosure not in csv_text:
            errors.append("phase2_geography_control: declared-review disclosure missing from manifest or risk sheet")
        for forbidden in ("Region Q official address", "fixture_search", "executor_run_001", "review_session_run_001"):
            if forbidden in payload or forbidden in csv_text:
                errors.append(f"phase2_geography_control: leaked internal search/session data {forbidden}")
        if "search_001" not in payload or "\"search_log_count\": 1" not in payload:
            errors.append("phase2_geography_control: manifest omitted internal SearchLog trace IDs/count")
        try:
            import openpyxl  # type: ignore
            xlsx_dir = directory / "phase2_geography_xlsx"
            xlsx = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(directory / "phase2_geography_control.json"), "--output-dir", str(xlsx_dir), "--mode", "standard", "--format", "xlsx"])
            if xlsx.returncode != 0:
                errors.append(f"phase2_geography_xlsx: export failed\n{xlsx.stdout}")
            else:
                workbook = openpyxl.load_workbook(xlsx_dir / "superleads_workbook.xlsx", read_only=True, data_only=True)
                cells = "\n".join(str(value) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True) for value in row if value is not None)
                for forbidden in ("Region Q official address", "fixture_search", "executor_run_001", "review_session_run_001"):
                    if forbidden in cells:
                        errors.append(f"phase2_geography_xlsx: leaked internal search/session data {forbidden}")
        except ImportError:
            pass

    full = copy.deepcopy(geography)
    full["briefs"][0]["evidence_depth"] = "full_review"
    _refresh_current_attestation(full)
    graph_path = directory / "phase2_full_unavailable.json"
    _write_graph(graph_path, full)
    full_audit = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(graph_path), "--delivery-status", "full_review_package", "--format", "json"])
    full_export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(directory / "phase2_full_unavailable"), "--mode", "full", "--format", "csv"])
    full_code = "full_review_unavailable_in_local_deployment"
    if full_audit.returncode == 0 or full_export.returncode == 0 or full_code not in full_audit.stdout or full_code not in full_export.stdout:
        errors.append(f"phase2_full_unavailable: full delivery was not fail-closed\n{full_audit.stdout}\n{full_export.stdout}")

    declared = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
    declared["briefs"][0]["evidence_depth"] = "full_review"
    subject = review_subject_hash(declared)
    declared["review_attestations"][0]["input_graph_hash"] = subject
    declared["review_attestations"][0]["reviewed_subject_hash"] = subject
    graph_path = directory / "phase2_declared_full.json"
    _write_graph(graph_path, declared)
    declared_audit = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(graph_path), "--delivery-status", "full_review_package", "--format", "json"])
    declared_export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(directory / "phase2_declared_full"), "--mode", "full", "--format", "csv"])
    if declared_audit.returncode == 0 or declared_export.returncode == 0 or "full_review_unavailable_in_local_deployment" not in declared_audit.stdout:
        errors.append(f"phase2_declared_full: declared separate session bypassed full block\n{declared_audit.stdout}\n{declared_export.stdout}")

    proxy = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
    proxy["observations"][0]["raw_excerpt"] = "Example Buyer contact +61 2 5555 0101. English service page at example.au."
    proxy["claims"][0]["typed_value"] = {"text": "Region Q"}
    proxy["claim_evidence"][0]["claim_field_anchors"]["claim_type"] = "Example Buyer contact"
    proxy["claim_evidence"][0]["claim_field_anchors"]["typed_value"] = "Region Q"
    proxy["review_attestations"][0]["input_graph_hash"] = review_subject_hash(proxy)
    proxy["review_attestations"][0]["reviewed_subject_hash"] = review_subject_hash(proxy)
    proxy_path = directory / "phase2_geography_proxy.json"
    _write_graph(proxy_path, proxy)
    proxy_result = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(proxy_path), "--delivery-status", "standard_development_list", "--format", "json"])
    if proxy_result.returncode == 0 or "geography_rule_support_not_formal_location" not in proxy_result.stdout:
        errors.append(f"phase2_geography_proxy: proxy signal bypassed geography evidence\n{proxy_result.stdout}")
    return errors


def _assert_target_geography_contract_required(directory: Path) -> list[str]:
    """Independently recreate the missing-contract geography delivery bypass."""
    errors: list[str] = []
    attack = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
    brief = attack["briefs"][0]
    contract = brief["customer_selection_contract"]
    contract.pop("geography_contract")
    contract["selection_requirements"][0].update({
        "rule_id": "sel_product_001",
        "user_statement": "Include organizations whose public page says it sells sample product.",
        "evidence_needed": "A same-Entity public page states sample product.",
        "search_hints": ["sample product seller"],
        "allowed_claim_types": ["product_match"],
        "evidence_markers": ["sample product"],
    })
    plan = attack["plans"][0]
    plan["selection_requirement_ids"] = ["sel_product_001"]
    plan["positive_query_groups"] = ["product_fit"]
    plan.pop("geography_query_group_ids", None)
    plan["query_groups"][0].update({
        "group_id": "product_fit",
        "query_purpose": "Find public product evidence before formal assessment.",
        "targeting_rule_ids": ["sel_product_001"],
        "queries": ["sample product seller"],
    })
    attack["candidates"][0].update({"discovery_method": "other", "search_log_id": None})
    attack["search_logs"] = []
    attack["observations"][0]["raw_excerpt"] = "Example Buyer sells sample product. Example Buyer Contact: sales@example.com"
    attack["claims"][0].update({"claim_type": "product_match", "predicate": "sells", "typed_value": {"text": "sample product"}})
    attack["claim_evidence"][0]["claim_field_anchors"].update({
        "predicate": "sells",
        "claim_type": "Example Buyer sells sample product",
        "typed_value": "sample product",
    })
    evaluation = attack["scope_decisions"][0]["rule_evaluations"][0]
    evaluation["rule_id"] = "sel_product_001"
    evaluation["claim_classifications"][0].update({"matched_marker": "sample product", "reason": "The public source contains the product marker."})
    evaluation["reason"] = "The opened public source states sample product."
    attack["scope_decisions"][0]["decision_summary"] = "Opened public source supports the product rule."
    _refresh_current_attestation(attack)
    validate, audit, export = _formal_results(attack, directory, "target_geography_contract_missing")
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0 or "geography_contract_required_for_target" not in result.stdout:
            errors.append(f"target_geography_contract_missing: {label} did not fail closed\n{result.stdout}")

    global_graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_global_target_without_geography_contract.json")
    validate, audit, export = _formal_results(global_graph, directory, "global_target_without_geography_contract")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        errors.append(f"global_target_without_geography_contract: null target unexpectedly required geography contract\n{validate.stdout}\n{audit.stdout}\n{export.stdout}")
    return errors


def _assert_legacy_review_fields_rejected(directory: Path) -> list[str]:
    """Schema must reject retired review-field shapes without keeping fixtures."""
    errors: list[str] = []
    field_name = "host" + "_identity_attestation"
    provenance_value = "host" + "_verified"
    legacy_signature_field = "signature" + "_base64"
    for name, mutate in (
        ("legacy_run_field", lambda graph: graph["runs"][0].update({field_name: {legacy_signature_field: "x"}})),
        ("legacy_attestation_value", lambda graph: graph["review_attestations"][0].update({"provenance_level": provenance_value})),
    ):
        graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
        mutate(graph)
        validate, audit, export = _formal_results(graph, directory, name)
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or "schema_validation_failed" not in result.stdout:
                errors.append(f"{name}: {label} did not schema-block retired review data\n{result.stdout}")

    for name, collection in (("legacy_audit_value", "audits"), ("legacy_manifest_value", "delivery_manifests")):
        graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
        graph[collection].append({"review_provenance_level": provenance_value})
        validate, audit, export = _formal_results(graph, directory, name)
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or "schema_validation_failed" not in result.stdout:
                errors.append(f"{name}: {label} did not schema-block retired provenance\n{result.stdout}")
    return errors


def _assert_review_attestation_coverage_and_disclosures(directory: Path) -> list[str]:
    """Exercise independent-session, Entity coverage, and stored disclosure gates."""
    errors: list[str] = []
    base = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_geography_searchlog_standard.json")
    variants = (
        (
            "attestation_same_session_only",
            lambda graph: graph["review_attestations"][0].update({"reviewer_session_id": graph["runs"][0]["execution_session_id"]}),
            "review_attestation_reviewer_session_not_independent",
        ),
        (
            "attestation_entity_coverage_only",
            lambda graph: graph["review_attestations"][0].update({"reviewed_entity_ids": []}),
            "review_attestation_entity_coverage_missing",
        ),
    )
    for name, mutate, expected_code in variants:
        graph = copy.deepcopy(base)
        mutate(graph)
        _refresh_current_attestation(graph)
        validate, audit, export = _formal_results(graph, directory, name)
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or expected_code not in result.stdout:
                errors.append(f"{name}: {label} did not enforce {expected_code}\n{result.stdout}")

    graph = copy.deepcopy(base)
    current_hash = graph_hash(graph)
    graph["audits"].append({
        "audit_id": "audit_disclosure_001",
        "audited_at": "2026-07-17T00:00:00Z",
        "research_graph_hash": current_hash,
        "audit_graph_hash": current_hash,
        "review_cycle_id": "review_run_001",
        "review_attestation_id": "att_run_001",
        "reviewed_subject_hash": current_hash,
        "review_provenance_level": "declared_separate_session",
        "audit_status": "passed",
        "delivery_status": "standard_development_list",
        "allowed_delivery_statuses": ["initial_lead_list", "standard_development_list"],
        "disclosure_required": False,
        "ok": True,
        "issue_count": 0,
        "issues": [],
    })
    graph["delivery_manifests"].append({
        "delivery_manifest_id": "manifest_disclosure_001",
        "run_id": "run_001",
        "brief_id": "brief_001",
        "plan_id": "plan_001",
        "audit_id": "audit_disclosure_001",
        "audit_graph_hash": current_hash,
        "research_graph_hash": current_hash,
        "review_cycle_id": "review_run_001",
        "review_attestation_id": "att_run_001",
        "reviewed_subject_hash": current_hash,
        "review_provenance_level": "declared_separate_session",
        "generated_at": "2026-07-17T00:00:00Z",
        "delivery_status": "standard_development_list",
        "output_mode": "standard",
        "exported_entity_ids": ["ent_001"],
        "exported_contact_ids": ["contact_001"],
        "exported_contact_claim_ids": ["cc_001"],
        "exported_assessment_ids": ["assess_001"],
        "output_files": [],
        "warnings": [],
        "disclosures": [],
    })
    validate, audit, export = _formal_results(graph, directory, "stored_declared_disclosure_missing")
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0 or "audit_disclosure_required_missing" not in result.stdout or "manifest_declared_review_disclosure_missing" not in result.stdout:
            errors.append(f"stored_declared_disclosure_missing: {label} did not enforce stored disclosure metadata\n{result.stdout}")
    return errors


def _assert_hold_value_is_not_exported(directory: Path) -> list[str]:
    graph = _base()
    graph["contact_claims"][0]["export_status"] = "hold_inferred"
    graph["contact_claims"][0]["user_status"] = "不可导出"
    _append_unassigned(graph)
    _refresh_current_attestation(graph)
    validate, audit, export = _formal_results(graph, directory, "hold_value_filtered")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        return [f"hold_value_filtered: expected export to pass\n{audit.stdout}\n{export.stdout}"]
    exported = "\n".join(path.read_text(encoding="utf-8-sig") for path in (directory / "hold_value_filtered").glob("*.csv"))
    return ["hold_value_filtered: hold contact value leaked into export"] if "sales@example.com" in exported else []


def _assert_historical_review_cannot_approve(directory: Path) -> list[str]:
    graph = _base()
    graph["observations"][0]["run_id"] = "run_001"
    graph["runs"].append({
        "run_id": "run_002",
        "brief_id": "brief_001",
        "plan_id": "plan_001",
        "created_at": "2026-01-02T00:00:00Z",
        "review_cycle_id": "review_run_002",
        "status": "checked",
        "platform": "fixture",
    })
    validate, audit, export = _formal_results(graph, directory, "historical_review_scope")
    errors = []
    if validate.returncode != 0:
        errors.append(f"historical_review_scope: graph should remain structurally valid\n{validate.stdout}")
    if audit.returncode == 0 or export.returncode == 0:
        errors.append(f"historical_review_scope: historical independent review approved current Run\n{audit.stdout}\n{export.stdout}")
    return errors


def _assert_historical_assessment_cannot_be_reused(directory: Path) -> list[str]:
    graph = _base()
    graph["observations"][0]["run_id"] = "run_001"
    graph["runs"].append({
        "run_id": "run_002",
        "brief_id": "brief_001",
        "plan_id": "plan_001",
        "created_at": "2026-01-02T00:00:00Z",
        "review_cycle_id": "review_run_002",
        "review_mode": "independent",
        "status": "checked",
        "platform": "fixture",
    })
    validate, audit, export = _formal_results(graph, directory, "historical_assessment_scope")
    errors = []
    if validate.returncode != 0:
        errors.append(f"historical_assessment_scope: graph should remain structurally valid\n{validate.stdout}")
    if audit.returncode == 0 or export.returncode == 0:
        errors.append(f"historical_assessment_scope: current Run reused an Assessment from another Run\n{audit.stdout}\n{export.stdout}")
    return errors


def _assert_formal_exception_bindings(directory: Path) -> list[str]:
    errors: list[str] = []
    for mode, expected_code in (
        ("single_company_analysis", "single_company_target_missing"),
        ("existing_table_enrichment", "existing_table_binding_missing"),
    ):
        graph = _base()
        graph["briefs"][0]["task_mode"] = mode
        graph["briefs"][0].pop("customer_selection_contract", None)
        graph["scope_decisions"] = []
        graph["assessments"][0].pop("scope_decision_id", None)
        validate, audit, export = _formal_results(graph, directory, f"{mode}_missing_binding")
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or expected_code not in result.stdout:
                errors.append(f"{mode}_missing_binding: {label} did not block the unbound exception\n{result.stdout}")
    for mode, expected_code in (
        ("single_company_analysis", "single_company_assessment_outside_target"),
        ("existing_table_enrichment", "existing_table_assessment_outside_bound_input"),
    ):
        graph = _base()
        graph["briefs"][0]["task_mode"] = mode
        graph["briefs"][0].pop("customer_selection_contract", None)
        graph["scope_decisions"] = []
        graph["assessments"][0].pop("scope_decision_id", None)
        if mode == "single_company_analysis":
            graph["briefs"][0]["single_company_target"] = {
                "user_statement": "Analyze Example Buyer.", "company_name": "Example Buyer",
                "website_or_domain": "https://example.com", "source_id": None, "entity_literal": None, "resolved_entity_id": "ent_001",
            }
        else:
            graph["sources"].append({
                "source_id": "src_table_001", "publisher_relation": "unknown", "provenance": "user_provided",
                "material_role": "user_business_dataset", "medium": "spreadsheet", "access_boundary": "user_supplied",
                "owner_hint": "existing table", "artifact_sha256": "f" * 64, "artifact_name": "existing.xlsx",
                "artifact_media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            })
            graph["observations"].append({
                "observation_id": "obs_table_001", "source_id": "src_table_001", "candidate_id": None, "entity_id": "ent_001",
                "capability": "document.extract", "concrete_tool": "fixture", "observed_at": "2026-01-01T00:00:00Z",
                "access_status": "ok", "http_status": None, "title": "existing table", "raw_excerpt": "Example Buyer",
                "page_or_dom_locator": "Rows!A2", "content_hash": "table_001", "extraction_method": "fixture",
                "tool_version": "fixture", "language": "en", "translation_status": "original",
                "derived_from_observation_id": None, "snapshot_ref": "artifact:sha256:" + "f" * 64 + "#sheet=Rows&range=A2",
            })
            graph["briefs"][0]["existing_table_binding"] = {
                "source_id": "src_table_001",
                "entity_bindings": [{"entity_id": "ent_001", "observation_id": "obs_table_001", "row_or_cell_locator": "Rows!A2", "entity_literal": "Example Buyer"}],
            }
        graph["entities"].append({"entity_id": "ent_002", "name": "Unbound Buyer", "website": "https://unbound.example"})
        graph["assessments"][0]["entity_id"] = "ent_002"
        validate, audit, export = _formal_results(graph, directory, f"{mode}_outside_binding")
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or expected_code not in result.stdout:
                errors.append(f"{mode}_outside_binding: {label} did not block the unbound Entity\n{result.stdout}")
    return errors


def _assert_identity_literal_bindings(directory: Path) -> list[str]:
    errors: list[str] = []
    graph = _base()
    graph["briefs"][0]["task_mode"] = "single_company_analysis"
    graph["briefs"][0].pop("customer_selection_contract", None)
    graph["scope_decisions"] = []
    graph["assessments"][0].pop("scope_decision_id", None)
    graph["briefs"][0]["single_company_target"] = {
        "user_statement": "Analyze the supplied company.", "company_name": "Unrelated Buyer",
        "website_or_domain": "https://example.com", "source_id": None, "entity_literal": None,
        "resolved_entity_id": "ent_001",
    }
    validate, audit, export = _formal_results(graph, directory, "single_company_conflicting_identifiers")
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0 or "single_company_target_identifier_conflict" not in result.stdout:
            errors.append(f"single_company_conflicting_identifiers: {label} did not block conflicting identifiers\n{result.stdout}")

    graph = _base()
    graph["briefs"][0]["task_mode"] = "existing_table_enrichment"
    graph["briefs"][0].pop("customer_selection_contract", None)
    graph["scope_decisions"] = []
    graph["assessments"][0].pop("scope_decision_id", None)
    graph["sources"].append({
        "source_id": "src_table_literal_001", "publisher_relation": "unknown", "provenance": "user_provided",
        "material_role": "user_business_dataset", "medium": "spreadsheet", "access_boundary": "user_supplied",
        "owner_hint": "existing table", "artifact_sha256": "e" * 64, "artifact_name": "existing.xlsx",
        "artifact_media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })
    graph["observations"].append({
        "observation_id": "obs_table_literal_001", "source_id": "src_table_literal_001", "candidate_id": None, "entity_id": "ent_001",
        "capability": "document.extract", "concrete_tool": "fixture", "observed_at": "2026-01-01T00:00:00Z",
        "access_status": "ok", "http_status": None, "title": "existing table", "raw_excerpt": "Unrelated Buyer",
        "page_or_dom_locator": "Rows!A2", "content_hash": "table_literal_001", "extraction_method": "fixture",
        "tool_version": "fixture", "language": "en", "translation_status": "original",
        "derived_from_observation_id": None, "snapshot_ref": "artifact:sha256:" + "e" * 64 + "#sheet=Rows&range=A2",
    })
    graph["briefs"][0]["existing_table_binding"] = {
        "source_id": "src_table_literal_001",
        "entity_bindings": [{"entity_id": "ent_001", "observation_id": "obs_table_literal_001", "row_or_cell_locator": "Rows!A2", "entity_literal": "Unrelated Buyer"}],
    }
    validate, audit, export = _formal_results(graph, directory, "existing_table_wrong_row_identity")
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0 or "existing_table_binding_literal_entity_mismatch" not in result.stdout:
            errors.append(f"existing_table_wrong_row_identity: {label} did not block mismatched row identity\n{result.stdout}")
    return errors


def _background_graph(*, resolved: bool = False) -> dict[str, Any]:
    target: dict[str, Any] = {
        "user_statement": "请对 Chilly's 品牌做客户背调。",
        "anchors": [
            {
                "anchor_id": "anchor_chillys_company_001",
                "kind": "company_name",
                "literal": "Chilly's",
                "candidate_id": None,
                "source_id": None,
            }
        ],
        "subject_resolution_status": "unresolved",
        "primary_subject_entity_id": None,
        "resolution_observation_ids": [],
    }
    graph: dict[str, Any] = {
        "runs": [{
            "run_id": "run_background_chillys_001",
            "status": "scoped",
            "created_at": "2026-07-19T00:00:00Z",
            "platform": "fixture",
            "brief_id": "brief_background_chillys_001",
            "review_mode": "not_run",
        }],
        "briefs": [{
            "brief_id": "brief_background_chillys_001",
            "task_mode": "customer_background_research",
            "output_mode": "客户背调报告",
            "contact_detail_level": "standard",
            "background_research_target": target,
        }],
        "plans": [],
        "candidates": [],
    }
    if not resolved:
        return graph

    target.update({
        "anchors": [
            *target["anchors"],
            {
                "anchor_id": "anchor_chillys_brand_001",
                "kind": "brand_name",
                "literal": "Chilly's",
                "candidate_id": None,
                "source_id": None,
            },
            {
                "anchor_id": "anchor_chillys_website_001",
                "kind": "website_or_domain",
                "literal": "https://chillys.example",
                "candidate_id": None,
                "source_id": None,
            },
        ],
        "subject_resolution_status": "resolved",
        "primary_subject_entity_id": "ent_chillys_legal_001",
        "resolution_observation_ids": ["obs_chillys_legal_001"],
    })
    graph.update({
        "entities": [{
            "entity_id": "ent_chillys_legal_001",
            "name": "Chilly's Bottles Limited",
            "legal_name": "Chilly's Bottles Limited",
            "website": "https://chillys.example",
        }],
        "sources": [{
            "source_id": "src_chillys_legal_001",
            "canonical_url": "https://chillys.example/legal",
            "final_url": "https://chillys.example/legal",
            "publisher_relation": "first_party",
            "provenance": "discovered_public",
            "medium": "website",
            "access_boundary": "public",
        }],
        "observations": [{
            "observation_id": "obs_chillys_legal_001",
            "source_id": "src_chillys_legal_001",
            "candidate_id": None,
            "entity_id": "ent_chillys_legal_001",
            "capability": "source.open",
            "concrete_tool": "fixture",
            "observed_at": "2026-07-19T00:00:00Z",
            "access_status": "ok",
            "http_status": 200,
            "title": "Chilly's legal information",
            "raw_excerpt": "Chilly's Bottles Limited is the legal entity named on this page.",
            "page_or_dom_locator": "main",
            "content_hash": "fixture_chillys_legal_001",
            "extraction_method": "fixture",
            "tool_version": "fixture",
            "language": "en",
            "translation_status": "original",
            "derived_from_observation_id": None,
            "snapshot_ref": None,
        }],
    })
    return graph


def _background_validate(graph: dict[str, Any], directory: Path, name: str) -> subprocess.CompletedProcess[str]:
    graph_path = directory / f"{name}.json"
    _write_graph(graph_path, graph)
    return _run([sys.executable, "-B", str(SCRIPTS / "validate_research_graph.py"), str(graph_path), "--format", "json"])


def _background_report_graph() -> dict[str, Any]:
    """A Chilly's-style graph with only the smallest evidence-backed report surface."""
    graph = _background_graph(resolved=True)
    target = graph["briefs"][0]["background_research_target"]
    target["anchors"].append({
        "anchor_id": "anchor_chillys_material_001",
        "kind": "user_material",
        "literal": "用户提供的 Similarweb 摘要",
        "candidate_id": None,
        "source_id": "src_chillys_material_001",
    })
    graph["entities"].append({
        "entity_id": "ent_chillys_brand_001",
        "name": "Chilly Brand IP Holdings",
        "legal_name": None,
        "website": "https://chillys.example",
    })
    graph["sources"].extend([
        {
            "source_id": "src_chillys_business_001",
            "canonical_url": "https://chillys.example/about",
            "final_url": "https://chillys.example/about",
            "publisher_relation": "first_party",
            "provenance": "discovered_public",
            "medium": "website",
            "access_boundary": "public",
        },
        {
            "source_id": "src_chillys_restricted_001",
            "canonical_url": "https://restricted.example/chillys",
            "final_url": "https://restricted.example/chillys",
            "publisher_relation": "third_party",
            "provenance": "discovered_public",
            "medium": "website",
            "access_boundary": "public",
        },
        {
            "source_id": "src_chillys_material_001",
            "publisher_relation": "unknown",
            "provenance": "manual_input",
            "material_role": "user_authored_note",
            "medium": "document",
            "access_boundary": "user_supplied",
        },
        {
            "source_id": "src_unrelated_bulk_001",
            "canonical_url": "https://unrelated.example/about",
            "final_url": "https://unrelated.example/about",
            "publisher_relation": "first_party",
            "provenance": "discovered_public",
            "medium": "website",
            "access_boundary": "public",
        },
    ])
    graph["observations"].extend([
        {
            "observation_id": "obs_chillys_business_001",
            "source_id": "src_chillys_business_001",
            "candidate_id": None,
            "entity_id": "ent_chillys_legal_001",
            "capability": "source.open",
            "concrete_tool": "fixture",
            "observed_at": "2026-07-19T00:00:00Z",
            "access_status": "ok",
            "http_status": 200,
            "title": "Chilly's about and contact",
            "raw_excerpt": "Chilly's Bottles Limited offers reusable bottles. Chilly's Bottles Limited operates as a wholesale brand. Chilly's Bottles Limited lists wholesale@chillys.example. Founder Jane Example.",
            "page_or_dom_locator": "main",
            "content_hash": "fixture_chillys_business_001",
            "extraction_method": "fixture",
            "tool_version": "fixture",
            "language": "en",
            "translation_status": "original",
            "derived_from_observation_id": None,
            "snapshot_ref": None,
        },
        {
            "observation_id": "obs_chillys_restricted_001",
            "source_id": "src_chillys_restricted_001",
            "candidate_id": None,
            "entity_id": "ent_chillys_legal_001",
            "capability": "source.open",
            "concrete_tool": "fixture",
            "observed_at": "2026-07-19T00:00:00Z",
            "access_status": "login_wall",
            "http_status": 403,
            "title": "Restricted third-party profile",
            "raw_excerpt": None,
            "page_or_dom_locator": None,
            "content_hash": None,
            "extraction_method": "fixture",
            "tool_version": "fixture",
            "language": "en",
            "translation_status": "original",
            "derived_from_observation_id": None,
            "snapshot_ref": None,
        },
        {
            "observation_id": "obs_unrelated_bulk_001",
            "source_id": "src_unrelated_bulk_001",
            "candidate_id": "cand_unrelated_bulk_001",
            "entity_id": "ent_unrelated_bulk_001",
            "capability": "source.open",
            "concrete_tool": "fixture",
            "observed_at": "2026-07-19T00:00:00Z",
            "access_status": "ok",
            "http_status": 200,
            "title": "Unrelated bulk company",
            "raw_excerpt": "Unrelated Bulk Limited offers unrelated products.",
            "page_or_dom_locator": "main",
            "content_hash": "fixture_unrelated_bulk_001",
            "extraction_method": "fixture",
            "tool_version": "fixture",
            "language": "en",
            "translation_status": "original",
            "derived_from_observation_id": None,
            "snapshot_ref": None,
        },
    ])
    graph["claims"] = [
        {
            "claim_id": "claim_chillys_identity_001",
            "entity_id": "ent_chillys_legal_001",
            "claim_type": "company_identity",
            "subject": "Chilly's Bottles Limited",
            "predicate": "is",
            "typed_value": {"text": "Chilly's Bottles Limited"},
            "as_of": "2026-07-19",
            "claim_scope": "current_source",
            "support_status": "supported",
            "contradiction_status": "none",
        },
        {
            "claim_id": "claim_chillys_product_001",
            "entity_id": "ent_chillys_legal_001",
            "claim_type": "product_match",
            "subject": "Chilly's Bottles Limited",
            "predicate": "offers",
            "typed_value": {"text": "reusable bottles"},
            "as_of": "2026-07-19",
            "claim_scope": "current_source",
            "support_status": "supported",
            "contradiction_status": "none",
        },
        {
            "claim_id": "claim_chillys_channel_001",
            "entity_id": "ent_chillys_legal_001",
            "claim_type": "channel_role",
            "subject": "Chilly's Bottles Limited",
            "predicate": "operates_as",
            "typed_value": {"text": "wholesale brand"},
            "as_of": "2026-07-19",
            "claim_scope": "current_source",
            "support_status": "supported",
            "contradiction_status": "none",
        },
    ]
    graph["claim_evidence"] = [
        {
            "claim_evidence_id": "ce_chillys_identity_001",
            "claim_id": "claim_chillys_identity_001",
            "observation_id": "obs_chillys_legal_001",
            "relation": "supports",
            "directness": "direct",
            "source_authority": "first_party",
            "independence_group": "chillys.example",
            "freshness": "current",
            "excerpt_pointer": "main",
            "claim_field_anchors": {
                "subject": "Chilly's Bottles Limited",
                "predicate": "is",
                "claim_type": "legal entity",
                "typed_value": "Chilly's Bottles Limited",
            },
        },
        {
            "claim_evidence_id": "ce_chillys_product_001",
            "claim_id": "claim_chillys_product_001",
            "observation_id": "obs_chillys_business_001",
            "relation": "supports",
            "directness": "direct",
            "source_authority": "first_party",
            "independence_group": "chillys.example",
            "freshness": "current",
            "excerpt_pointer": "main",
            "claim_field_anchors": {
                "subject": "Chilly's Bottles Limited",
                "predicate": "offers",
                "claim_type": "reusable bottles",
                "typed_value": "reusable bottles",
            },
        },
        {
            "claim_evidence_id": "ce_chillys_channel_001",
            "claim_id": "claim_chillys_channel_001",
            "observation_id": "obs_chillys_business_001",
            "relation": "supports",
            "directness": "direct",
            "source_authority": "first_party",
            "independence_group": "chillys.example",
            "freshness": "current",
            "excerpt_pointer": "main",
            "claim_field_anchors": {
                "subject": "Chilly's Bottles Limited",
                "predicate": "operates as",
                "claim_type": "wholesale brand",
                "typed_value": "wholesale brand",
            },
        },
    ]
    graph["entity_relationships"] = [{
        "entity_relationship_id": "rel_chillys_legal_brand_001",
        "source_entity_id": "ent_chillys_legal_001",
        "target_entity_id": "ent_chillys_brand_001",
        "relationship_type": "legal_entity_of",
        "resolution_status": "contextual",
        "confidence": "high",
        "rationale": "Legal page identifies the legal entity for the Chilly's brand.",
        "evidence_claim_ids": ["claim_chillys_identity_001"],
        "evidence_observation_ids": ["obs_chillys_legal_001"],
    }]
    graph["contact_points"] = [
        {
            "contact_id": "contact_chillys_ready_001",
            "contact_type": "department_email",
            "normalized_value": "wholesale@chillys.example",
            "source_literal": "wholesale@chillys.example",
            "source_observation_id": "obs_chillys_business_001",
            "source_type": "website",
            "visibility_status": "public",
            "last_seen_at": "2026-07-19T00:00:00Z",
            "verification_status": "not_verified",
        },
        {
            "contact_id": "contact_chillys_hold_001",
            "contact_type": "email",
            "normalized_value": "hidden@chillys.example",
            "source_literal": "hidden@chillys.example",
            "source_observation_id": "obs_chillys_business_001",
            "source_type": "website",
            "visibility_status": "public",
            "last_seen_at": "2026-07-19T00:00:00Z",
            "verification_status": "not_verified",
        },
        {
            "contact_id": "contact_chillys_unassigned_001",
            "contact_type": "email",
            "normalized_value": "unassigned@chillys.example",
            "source_literal": "unassigned@chillys.example",
            "source_observation_id": "obs_chillys_business_001",
            "source_type": "website",
            "visibility_status": "public",
            "last_seen_at": "2026-07-19T00:00:00Z",
            "verification_status": "not_verified",
        },
    ]
    graph["observations"][1]["raw_excerpt"] = "Chilly's Bottles Limited offers reusable bottles. Chilly's Bottles Limited operates as a wholesale brand. Chilly's Bottles Limited lists wholesale@chillys.example, hidden@chillys.example, and unassigned@chillys.example. Founder Jane Example."
    graph["contact_claims"] = [
        {
            "contact_claim_id": "cc_chillys_ready_001",
            "contact_id": "contact_chillys_ready_001",
            "entity_id": "ent_chillys_legal_001",
            "person_id": None,
            "person_name": "Jane Example",
            "job_title": "Founder",
            "department": None,
            "relationship_type": "company_general",
            "association_observation_id": "obs_chillys_business_001",
            "association_claim_evidence_ids": [],
            "source_context": "first-party wholesale contact",
            "association_evidence_text": "Chilly's Bottles Limited lists wholesale@chillys.example, hidden@chillys.example, and unassigned@chillys.example. Founder Jane Example.",
            "association_locator": "main",
            "association_confidence": "high",
            "is_role_based": True,
            "is_personal_business": False,
            "export_status": "ready",
            "user_status": "可直接使用",
            "manual_check_note": None,
        },
        {
            "contact_claim_id": "cc_chillys_hold_001",
            "contact_id": "contact_chillys_hold_001",
            "entity_id": "ent_chillys_legal_001",
            "person_id": None,
            "person_name": None,
            "job_title": None,
            "department": None,
            "relationship_type": "company_general",
            "association_observation_id": "obs_chillys_business_001",
            "association_claim_evidence_ids": [],
            "source_context": "unverified contact value",
            "association_evidence_text": "Chilly's Bottles Limited lists wholesale@chillys.example, hidden@chillys.example, and unassigned@chillys.example.",
            "association_locator": "main",
            "association_confidence": "low",
            "is_role_based": False,
            "is_personal_business": False,
            "export_status": "hold_inferred",
            "user_status": "不可导出",
            "manual_check_note": "Do not expose before public verification.",
        },
    ]
    graph["unassigned_contact_leads"] = [{
        "unassigned_contact_lead_id": "unassigned_chillys_001",
        "contact_id": "contact_chillys_unassigned_001",
        "reason": "Public contact value has no confirmed entity association.",
        "suggested_manual_check": "Confirm the page owner before use.",
    }]
    graph["contact_points"].append({
        "contact_id": "contact_chillys_manual_person_001",
        "contact_type": "person_name",
        "normalized_value": "Manual Person",
        "source_literal": "Manual Person",
        "source_observation_id": "obs_chillys_business_001",
        "source_type": "website",
        "visibility_status": "public",
        "last_seen_at": "2026-07-19T00:00:00Z",
        "verification_status": "not_verified",
    })
    graph["contact_claims"].append({
        "contact_claim_id": "cc_chillys_manual_person_001",
        "contact_id": "contact_chillys_manual_person_001",
        "entity_id": "ent_chillys_legal_001",
        "person_id": None,
        "person_name": "Manual Person",
        "job_title": "Owner",
        "department": None,
        "relationship_type": "public_person_clue",
        "association_observation_id": "obs_chillys_business_001",
        "association_claim_evidence_ids": [],
        "source_context": "public person clue requiring manual association review",
        "association_evidence_text": "Manual Person Owner",
        "association_locator": "main",
        "association_confidence": "low",
        "is_role_based": False,
        "is_personal_business": False,
        "export_status": "needs_manual_association_review",
        "user_status": "待确认归属",
        "manual_check_note": "Manual Person is not confirmed as a purchasing contact.",
    })
    graph["observations"][1]["raw_excerpt"] += " Manual Person Owner."
    graph["hypotheses"] = [{
        "hypothesis_id": "hyp_chillys_wholesale_001",
        "entity_id": "ent_chillys_legal_001",
        "basis_claim_ids": ["claim_chillys_product_001", "claim_chillys_channel_001"],
        "basis_contact_claim_ids": ["cc_chillys_ready_001"],
        "hypothesis_text": "可围绕已观察到的可重复使用水瓶与批发品牌角色准备沟通角度。",
        "unknowns": ["批发入口是否仍面向新的供应商待确认"],
        "suggested_action": "根据公开业务信息准备初步材料。",
        "next_verification_action": "确认当前 wholesale 或 vendor 入口。",
        "expires_at": None,
        "risk_notes": ["不代表采购需求或采购职责已确认"],
    }]
    graph["entities"].append({
        "entity_id": "ent_unrelated_bulk_001",
        "name": "Unrelated Bulk Limited",
        "legal_name": None,
        "website": "https://unrelated.example",
    })
    graph["candidates"].append({
        "candidate_id": "cand_unrelated_bulk_001",
        "name": "Unrelated Bulk Limited",
        "company_name": "Unrelated Bulk Limited",
        "entity_id": "ent_unrelated_bulk_001",
    })
    graph["audits"] = [{"audit_id": "old_audit_unrelated_001", "delivery_status": "standard_development_list"}]
    graph["delivery_manifests"] = [{"delivery_manifest_id": "old_manifest_unrelated_001", "delivery_status": "standard_development_list"}]
    return graph


def _background_export(graph: dict[str, Any], directory: Path, name: str, output_format: str = "csv", manifest: bool = False) -> tuple[subprocess.CompletedProcess[str], Path]:
    graph_path = directory / f"{name}.json"
    output_dir = directory / name
    _write_graph(graph_path, graph)
    command = [sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(output_dir), "--mode", "background", "--format", output_format]
    if manifest:
        command.extend(["--manifest", str(output_dir / "manifest.json")])
    return _run(command), output_dir


def _assert_background_report_export(directory: Path) -> list[str]:
    errors: list[str] = []
    unresolved, unresolved_dir = _background_export(_background_graph(), directory, "background_report_unresolved")
    if unresolved.returncode != 0:
        errors.append(f"background_report_unresolved: no-Entity/no-Plan draft did not export\n{unresolved.stdout}")
    elif "\"manifest\": null" not in unresolved.stdout:
        errors.append("background_report_unresolved: background export unexpectedly returned a manifest")

    graph = _background_report_graph()
    result, output_dir = _background_export(graph, directory, "background_report_chillys")
    if result.returncode != 0:
        return errors + [f"background_report_chillys: resolved report did not export\n{result.stdout}"]
    rendered = "\n".join(path.read_text(encoding="utf-8-sig") for path in output_dir.glob("*.csv"))
    required_sheets = {"背调报告.csv", "客户与研究锚点.csv", "主体与关系.csv", "产品、渠道与经营信号.csv", "公开联系入口与桥接候选.csv", "开发切入点候选.csv", "谈判前待确认问题.csv", "未确认线索与来源受限.csv", "证据包.csv"}
    present_sheets = {path.name for path in output_dir.glob("*.csv")}
    if required_sheets != present_sheets:
        errors.append(f"background_report_chillys: unexpected CSV sheets {sorted(present_sheets)}")
    for needle in ("Chilly's Bottles Limited", "Chilly's", "reusable bottles", "wholesale@chillys.example", "来源受限"):
        if needle not in rendered:
            errors.append(f"background_report_chillys: missing expected report content {needle}")
    for forbidden in ("hidden@chillys.example", "unassigned@chillys.example", "Manual Person", "Unrelated Bulk Limited", "old_audit_unrelated_001", "old_manifest_unrelated_001"):
        if forbidden in rendered:
            errors.append(f"background_report_chillys: leaked out-of-scope or hidden value {forbidden}")
    if "[已隐藏联系方式]" not in rendered:
        errors.append("background_report_chillys: hidden contact evidence was not redacted")

    bad_manifest, _ = _background_export(_background_graph(), directory, "background_report_manifest_rejected", manifest=True)
    if bad_manifest.returncode == 0 or "--manifest is not supported for --mode background" not in bad_manifest.stdout:
        errors.append(f"background_report_manifest_rejected: manifest was not rejected\n{bad_manifest.stdout}")

    mismatch = _background_graph()
    mismatch["briefs"][0]["output_mode"] = "发现候选池"
    bad_mode, _ = _background_export(mismatch, directory, "background_report_mode_mismatch")
    if bad_mode.returncode == 0 or "background_export_output_mode_mismatch" not in bad_mode.stdout:
        errors.append(f"background_report_mode_mismatch: mismatched output mode was not blocked\n{bad_mode.stdout}")

    task_mismatch = _background_graph()
    task_mismatch["briefs"][0]["task_mode"] = "single_company_analysis"
    bad_task, _ = _background_export(task_mismatch, directory, "background_report_task_mismatch")
    if bad_task.returncode == 0 or "background_export_task_mode_mismatch" not in bad_task.stdout:
        errors.append(f"background_report_task_mismatch: mismatched task mode was not blocked\n{bad_task.stdout}")

    auto, auto_dir = _background_export(_background_graph(), directory, "background_report_auto", output_format="auto")
    if auto.returncode != 0:
        errors.append(f"background_report_auto: auto format export failed\n{auto.stdout}")
    else:
        try:
            payload = json.loads(auto.stdout)
        except json.JSONDecodeError:
            errors.append("background_report_auto: output was not JSON")
        else:
            if payload.get("format") == "xlsx":
                try:
                    import openpyxl  # type: ignore
                    workbook = openpyxl.load_workbook(auto_dir / "superleads_background_report.xlsx", read_only=True, data_only=True)
                    if set(workbook.sheetnames) != {name[:-4] for name in required_sheets}:
                        errors.append(f"background_report_auto: unexpected XLSX sheet names {workbook.sheetnames}")
                except Exception as exc:
                    errors.append(f"background_report_auto: could not inspect XLSX output: {exc}")
            elif payload.get("format") == "csv":
                if {path.name for path in auto_dir.glob("*.csv")} != required_sheets:
                    errors.append("background_report_auto: CSV fallback lacks expected sheets")
            else:
                errors.append(f"background_report_auto: unknown chosen format {payload.get('format')}")
    return errors


def _assert_background_research_contract(directory: Path) -> list[str]:
    """Keep the Chilly's-derived background-research draft contract isolated from delivery paths."""
    errors: list[str] = []

    unresolved = _background_validate(_background_graph(), directory, "background_chillys_unresolved")
    if unresolved.returncode != 0:
        errors.append(f"background_chillys_unresolved: unresolved no-Plan/no-Entity draft failed\n{unresolved.stdout}")

    resolved = _background_validate(_background_graph(resolved=True), directory, "background_chillys_resolved")
    if resolved.returncode != 0:
        errors.append(f"background_chillys_resolved: sourced brand/website-to-legal-Entity draft failed\n{resolved.stdout}")

    variants: list[tuple[str, Callable[[dict[str, Any]], None], tuple[str, ...]]] = []

    def add(name: str, mutate: Callable[[dict[str, Any]], None], *codes: str) -> None:
        variants.append((name, mutate, codes))

    def missing_candidate(graph: dict[str, Any]) -> None:
        graph["briefs"][0]["background_research_target"]["anchors"][0] = {
            "anchor_id": "anchor_missing_candidate_001", "kind": "candidate_id", "literal": None,
            "candidate_id": "cand_missing_001", "source_id": None,
        }

    add("background_missing_candidate", missing_candidate, "background_candidate_anchor_missing")
    add("background_duplicate_anchor_id", lambda graph: graph["briefs"][0]["background_research_target"]["anchors"].append({
        "anchor_id": "anchor_chillys_company_001", "kind": "email", "literal": "research@chillys.example",
        "candidate_id": None, "source_id": None,
    }), "background_anchor_id_duplicate")

    def missing_material_source(graph: dict[str, Any]) -> None:
        graph["briefs"][0]["background_research_target"]["anchors"][0] = {
            "anchor_id": "anchor_missing_material_001", "kind": "user_material", "literal": "客户上传的线索",
            "candidate_id": None, "source_id": "src_missing_001",
        }

    add("background_missing_material_source", missing_material_source, "background_user_material_source_missing")

    def public_material_source(graph: dict[str, Any]) -> None:
        graph["sources"] = [{
            "source_id": "src_public_001", "canonical_url": "https://chillys.example/about", "final_url": "https://chillys.example/about",
            "publisher_relation": "first_party", "provenance": "discovered_public", "medium": "website", "access_boundary": "public",
        }]
        graph["briefs"][0]["background_research_target"]["anchors"][0] = {
            "anchor_id": "anchor_public_material_001", "kind": "user_material", "literal": "客户提供链接说明",
            "candidate_id": None, "source_id": "src_public_001",
        }

    add("background_public_material_source", public_material_source, "background_user_material_source_not_user_provided")

    def missing_entity(graph: dict[str, Any]) -> None:
        graph["briefs"][0]["background_research_target"]["primary_subject_entity_id"] = "ent_missing_001"

    add("background_missing_entity", missing_entity, "background_primary_subject_entity_missing", "background_resolution_observation_entity_mismatch")
    add("background_wrong_observation_entity", lambda graph: graph["observations"][0].update({"entity_id": "ent_other_001"}), "background_resolution_observation_entity_mismatch")
    add("background_empty_resolution_excerpt", lambda graph: graph["observations"][0].update({"raw_excerpt": ""}), "background_resolution_observation_excerpt_missing")
    add("background_restricted_resolution", lambda graph: graph["observations"][0].update({"access_status": "login_wall"}), "background_resolution_observation_access_restricted")
    add("background_search_result_resolution", lambda graph: graph["sources"][0].update({"medium": "search_result"}), "background_resolution_source_search_result")
    add("background_uncheckable_resolution_capability", lambda graph: graph["observations"][0].update({"capability": "company.enrich"}), "background_resolution_observation_capability_not_checkable")
    add("background_unsafe_website_anchor", lambda graph: graph["briefs"][0]["background_research_target"]["anchors"][0].update({"kind": "website_or_domain", "literal": "http://127.0.0.1/private"}), "background_website_or_domain_not_public")

    for name, mutate, codes in variants:
        graph = _background_graph(resolved=name.startswith("background_") and name not in {
            "background_missing_candidate", "background_missing_material_source", "background_public_material_source", "background_unsafe_website_anchor",
        })
        mutate(graph)
        result = _background_validate(graph, directory, name)
        missing_codes = [code for code in codes if code not in result.stdout]
        if result.returncode == 0 or missing_codes:
            errors.append(f"{name}: expected validation failure with {', '.join(codes)}\n{result.stdout}")

    return errors


def _assert_hold_free_text_is_redacted(directory: Path) -> list[str]:
    graph = _base()
    graph["contact_claims"][0]["export_status"] = "hold_inferred"
    graph["contact_claims"][0]["user_status"] = "不可导出"
    graph["assessments"][0]["missing_requirements"] = ["Do not expose sales@example.com"]
    _append_unassigned(graph)
    graph["unassigned_contact_leads"][0].update({
        "reason": "sales@example.com was inferred",
        "suggested_manual_check": "Verify sales@example.com only from a public page.",
    })
    _refresh_current_attestation(graph)
    validate, audit, export = _formal_results(graph, directory, "hold_free_text_redaction")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        return [f"hold_free_text_redaction: expected sanitized export to pass\n{audit.stdout}\n{export.stdout}"]
    output_dir = directory / "hold_free_text_redaction"
    exported = "\n".join(path.read_text(encoding="utf-8-sig") for path in output_dir.glob("*.csv"))
    exported += (output_dir / "manifest.json").read_text(encoding="utf-8")
    errors = ["hold_free_text_redaction: hold contact leaked through free text"] if "sales@example.com" in exported else []
    try:
        import openpyxl  # type: ignore
    except Exception:
        return errors
    graph_path = directory / "hold_free_text_redaction.json"
    xlsx_dir = directory / "hold_free_text_redaction_xlsx"
    xlsx = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(xlsx_dir), "--mode", "standard", "--format", "xlsx"])
    if xlsx.returncode != 0:
        errors.append(f"hold_free_text_redaction: XLSX export failed\n{xlsx.stdout}")
        return errors
    workbook = openpyxl.load_workbook(xlsx_dir / "superleads_workbook.xlsx", read_only=True, data_only=True)
    cells = "\n".join(str(cell) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None)
    if "sales@example.com" in cells:
        errors.append("hold_free_text_redaction: hold contact leaked into XLSX")
    return errors


def _assert_manual_contact_is_not_exported(directory: Path) -> list[str]:
    graph = _base()
    graph["contact_claims"][0].update({
        "export_status": "needs_manual_association_review",
        "user_status": "待确认归属",
        "manual_check_note": "Manual owner check required.",
    })
    _refresh_current_attestation(graph)
    validate, audit, export = _formal_results(graph, directory, "manual_contact_filtered")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        return [f"manual_contact_filtered: expected formal export to pass\n{audit.stdout}\n{export.stdout}"]
    contacts = (directory / "manual_contact_filtered" / "联系方式汇总.csv").read_text(encoding="utf-8-sig")
    return ["manual_contact_filtered: manual-association contact leaked into contact summary"] if "sales@example.com" in contacts else []


def _assert_standard_docs_include_source_links() -> list[str]:
    paths = [
        ROOT / "skills" / "exporting-lead-workbooks" / "SKILL.md",
        ROOT / "shared" / "references" / "output-schema.md",
    ]
    return [f"{path}: standard export documentation omits 官网与来源链接" for path in paths if "官网与来源链接" not in path.read_text(encoding="utf-8")]


def _assert_user_file_direct_pressure_tests(directory: Path) -> list[str]:
    errors: list[str] = []
    validate = _run([sys.executable, "-B", str(SCRIPTS / "validate_research_graph.py"), str(USER_FILE_PDF)])
    if validate.returncode != 0:
        errors.append(f"user_file_pdf_validate: expected pass\n{validate.stdout}")
    audit = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(USER_FILE_PDF), "--delivery-status", "standard_development_list", "--format", "json"])
    if audit.returncode != 0 or '"delivery_status": "standard_development_list"' not in audit.stdout:
        errors.append(f"user_file_pdf_audit: expected standard delivery\n{audit.stdout}")
    output = directory / "superleads-user-file-test.xlsx"
    export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(USER_FILE_SHEET), str(output), "--mode", "standard"])
    if export.returncode != 0 or not output.exists():
        errors.append(f"user_file_sheet_export: expected XLSX output\n{export.stdout}")
        return errors
    try:
        import openpyxl  # type: ignore
        workbook = openpyxl.load_workbook(output, read_only=True, data_only=True)
        cells = "\n".join(str(value) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True) for value in row if value is not None)
    except Exception as exc:
        return errors + [f"user_file_sheet_export: could not inspect XLSX: {exc}"]
    if "用户提供文件：客户名单.xlsx（工作表 Contacts，A2:F2）" not in cells:
        errors.append("user_file_sheet_export: workbook lacks safe spreadsheet source display")
    for forbidden in ("file://", "/home/", "C:\\", "bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb"):
        if forbidden in cells:
            errors.append(f"user_file_sheet_export: forbidden local/internal value leaked: {forbidden}")
    return errors


def _assert_material_role_direct_pressure_tests(directory: Path) -> list[str]:
    errors: list[str] = []
    for name, fixture, expected_label in (
        ("dataset", DATASET_CONTACT, "用户提供文件：历史客户表.xlsx（工作表 Contacts，A2:F2）"),
        ("correspondence", CORRESPONDENCE_CONTACT, "用户提供沟通记录：客户沟通.eml（章节 message-1）"),
    ):
        out_dir = directory / f"{name}_export"
        graph_path = directory / f"{name}_source_note.json"
        _write_graph(graph_path, _load_fixture_graph(fixture))
        export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(out_dir), "--mode", "standard", "--format", "csv", "--manifest", str(out_dir / "manifest.json")])
        if export.returncode != 0:
            errors.append(f"{name}_source_note_export: expected standard export\n{export.stdout}")
            continue
        text = "\n".join(path.read_text(encoding="utf-8-sig") for path in out_dir.glob("*.csv"))
        text += (out_dir / "manifest.json").read_text(encoding="utf-8")
        if expected_label not in text or "建议核查后使用" not in text:
            errors.append(f"{name}_source_note_export: missing source-note display or status")
        for forbidden in ("可直接使用", "material_role", "file://", "/home/", "C:\\", "bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb"):
            if forbidden in text:
                errors.append(f"{name}_source_note_export: forbidden value leaked: {forbidden}")
    visual_path = directory / "visual_candidate.json"
    visual_output = directory / "visual_export"
    _write_graph(visual_path, _load_fixture_graph(VISUAL_CANDIDATE))
    validate = _run([sys.executable, "-B", str(SCRIPTS / "validate_research_graph.py"), str(visual_path)])
    standard = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(visual_path), "--output-dir", str(visual_output), "--mode", "standard", "--format", "csv"])
    initial = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(visual_path), "--output-dir", str(visual_output), "--mode", "initial", "--format", "csv"])
    if validate.returncode != 0 or initial.returncode != 0 or standard.returncode == 0:
        errors.append(f"visual_candidate_delivery_levels: expected validate+initial pass and standard block\n{validate.stdout}\n{standard.stdout}\n{initial.stdout}")
    return errors


def _assert_inquiry_export_redaction(directory: Path) -> list[str]:
    graph_path = directory / "connected_inquiry.json"
    out_dir = directory / "connected_inquiry_export"
    _write_graph(graph_path, _load_fixture_graph(CONNECTED_INQUIRY))
    audit = _run([sys.executable, "-B", str(SCRIPTS / "audit_delivery.py"), str(graph_path), "--delivery-status", "inquiry_followup_queue", "--format", "json"])
    export = _run([sys.executable, "-B", str(SCRIPTS / "export_workbook.py"), str(graph_path), "--output-dir", str(out_dir), "--mode", "inquiry", "--format", "csv", "--manifest", str(out_dir / "manifest.json")])
    if audit.returncode != 0 or export.returncode != 0:
        return [f"inquiry_export: expected inquiry audit/export to pass\n{audit.stdout}\n{export.stdout}"]
    text = "\n".join(path.read_text(encoding="utf-8-sig") for path in out_dir.glob("*.csv"))
    text += (out_dir / "manifest.json").read_text(encoding="utf-8")
    errors = []
    if "邮件来信（2026-07-15）" not in text or "询盘待办" not in text:
        errors.append("inquiry_export: missing business-facing inquiry label")
    for forbidden in ("host-message-001", "host-thread-001", "mailbox_inbox_001", "dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd", "From: sales@example.com", "file://", "/home/", "C:\\"):
        if forbidden in text:
            errors.append(f"inquiry_export: leaked internal or full-mail value: {forbidden}")
    return errors


def _assert_mail_adapter_boundary(directory: Path) -> list[str]:
    output = directory / "mail_adapter_output.json"
    accepted = _run([sys.executable, "-B", str(SCRIPTS / "ingest_mail_read_result.py"), str(MAIL_ADAPTER_INPUT), "--output", str(output)])
    if accepted.returncode != 0 or not output.exists():
        return [f"mail_adapter_accept: expected bounded host result to normalize\n{accepted.stdout}"]
    payload = json.loads(output.read_text(encoding="utf-8"))
    text = json.dumps(payload, ensure_ascii=False)
    if "password" in text or "token" in text or "host-message-adapter-001" not in text:
        return ["mail_adapter_accept: unexpected normalized result boundary"]
    bad_input = dict(json.loads(MAIL_ADAPTER_INPUT.read_text(encoding="utf-8")))
    bad_input["access_token"] = "must-not-store"
    bad_path = directory / "mail_adapter_bad.json"
    _write_graph(bad_path, bad_input)
    rejected = _run([sys.executable, "-B", str(SCRIPTS / "ingest_mail_read_result.py"), str(bad_path), "--output", str(directory / "mail_adapter_bad_output.json")])
    return [] if rejected.returncode != 0 and "mail_input_forbidden_sensitive_field" in rejected.stdout else [f"mail_adapter_reject: sensitive host input was not blocked\n{rejected.stdout}"]


def _assert_platform_and_public_url_pressure_tests(directory: Path) -> list[str]:
    """Construct variants directly so coverage cannot depend on fixture names."""
    errors: list[str] = []
    platform_variants = ("curl ", "shell-curl", "codex_cli ", "CODEX_CLI")
    for index, platform in enumerate(platform_variants):
        graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_codex_shell_http_source_open_standard.json")
        graph["runs"][0]["platform"] = platform
        graph["runs"][0].pop("capability_adapter_report", None)
        validate, audit, export = _formal_results(graph, directory, f"platform_variant_{index}")
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or "run_platform_not_canonical" not in result.stdout:
                errors.append(f"platform_variant_{platform!r}: {label} did not block canonical-platform bypass\n{result.stdout}")

    for index, hostname in enumerate(("127.1", "2130706433", "0x7f000001")):
        graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_codex_shell_http_source_open_standard.json")
        url = f"http://{hostname}/private"
        graph["sources"][0].update({"canonical_url": url, "final_url": url})
        operation = graph["runs"][0]["capability_adapter_report"]["host_tools"]["shell_http"]["operations"]["open_source"]
        operation.update({"original_url": url, "final_url": url})
        validate, audit, export = _formal_results(graph, directory, f"shell_loopback_{index}")
        for label, result in (("validate", validate), ("audit", audit), ("export", export)):
            if result.returncode == 0 or "codex_shell_http_url_not_public" not in result.stdout:
                errors.append(f"shell_loopback_{hostname}: {label} did not block legacy IPv4 loopback\n{result.stdout}")

    generic = _base()
    generic["runs"][0]["platform"] = "hermes"
    generic["sources"][0].update({"canonical_url": "http://127.0.0.1/private", "final_url": "http://127.0.0.1/private"})
    validate, audit, export = _formal_results(generic, directory, "generic_loopback")
    for label, result in (("validate", validate), ("audit", audit), ("export", export)):
        if result.returncode == 0 or "public_source_url_not_safe" not in result.stdout:
            errors.append(f"generic_loopback: {label} did not block formal private URL\n{result.stdout}")

    public_shell = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_codex_shell_http_source_open_standard.json")
    validate, audit, export = _formal_results(public_shell, directory, "public_shell_control")
    if any(result.returncode != 0 for result in (validate, audit, export)):
        errors.append(f"public_shell_control: public HTTPS shell source no longer passes\n{validate.stdout}\n{audit.stdout}\n{export.stdout}")
    return errors


def _assert_default_discovery_export_filters_unsafe_urls() -> list[str]:
    """The workbook builder must filter unsafe links even when validation is bypassed."""
    graph = _load_fixture_graph(ROOT / "evals" / "fixtures" / "pass_default_discovery_candidate_pool.json")
    candidates = graph["candidates"]
    candidates[0]["source_url"] = "https://example.com/#/callback?access_token=secret"
    candidates[0]["discovery_refs"][0]["url"] = "https://example.com/#/auth?TOKEN=secret"
    candidates[0]["signal_summary"]["business_match"]["items"][0]["source_url"] = "https://example.com/#route?api-key=secret"
    candidates[0]["website"] = "http://127.0.0.1/private"
    candidates[1]["source_url"] = "http://192.168.10.2/private"
    candidates[1]["signal_summary"]["website_contact"]["items"][0]["source_url"] = "https://user:token@example.com/contact"
    candidates[1]["website"] = "file:///home/user/private"
    candidates[2]["website"] = "https://user:token@example.com/private"
    candidates[3]["website"] = "https://example.com/#/callback?access_token=secret"
    candidates[4]["website"] = "https://example.com/#/catalog"
    plain_domain_candidate = copy.deepcopy(candidates[4])
    plain_domain_candidate.update({
        "candidate_id": "cand_plain_domain_export_001",
        "name": "Plain Domain Discovery",
        "company_name": "Plain Domain Discovery",
        "website": "example.com",
        "domain": None,
    })
    candidates.append(plain_domain_candidate)
    graph["search_logs"][0]["result_refs"][0]["result_url"] = "https://example.com/#/return?sig=secret"

    sheets = build_initial_sheets(graph, {"issues": []})
    rendered = json.dumps(sheets, ensure_ascii=False)
    unsafe_values = (
        "https://example.com/#/callback?access_token=secret",
        "https://example.com/#/auth?TOKEN=secret",
        "https://example.com/#route?api-key=secret",
        "file:///home/user/private",
        "http://127.0.0.1/private",
        "http://192.168.10.2/private",
        "https://user:token@example.com/contact",
        "https://user:token@example.com/private",
        "https://example.com/#/return?sig=secret",
    )
    errors = [f"default_discovery_export_unsafe_url: leaked {value}" for value in unsafe_values if value in rendered]
    if "Alpha 官网产品页" not in rendered:
        errors.append("default_discovery_export_unsafe_url: filtered link removed its safe source label")
    if is_safe_public_http_url("https://example.com/#/callback?access_token=secret"):
        errors.append("default_discovery_export_unsafe_url: SPA fragment token passed shared URL validation")
    if not is_safe_public_http_url("https://example.com/#/catalog"):
        errors.append("default_discovery_export_unsafe_url: safe SPA fragment failed shared URL validation")
    for safe_value in ("https://example.com/#/catalog", "example.com"):
        if safe_value not in rendered:
            errors.append(f"default_discovery_export_unsafe_url: safe website/domain missing {safe_value}")
    return errors


def _assert_standard_export_filters_unsafe_entity_websites() -> list[str]:
    """Standard sheets must not leak Entity website/domain values if validation is bypassed."""
    errors: list[str] = []
    unsafe_values = (
        "http://127.0.0.1/private",
        "file:///home/user/private",
        "https://user:token@example.com/path",
        "https://example.com/?access_token=x",
        "127.0.0.1",
    )
    for index, value in enumerate(unsafe_values):
        graph = _base()
        entity = graph["entities"][0]
        if index == len(unsafe_values) - 1:
            entity["website"] = None
            entity["domain"] = value
        else:
            entity["website"] = value
        rendered = json.dumps(build_sheets(graph, {"issues": []}, "standard"), ensure_ascii=False)
        if value in rendered:
            errors.append(f"standard_export_unsafe_entity_website: leaked {value}")

    safe_url_graph = _base()
    safe_url_graph["entities"][0]["website"] = "https://example.com/path?id=123"
    safe_url_rendered = json.dumps(build_sheets(safe_url_graph, {"issues": []}, "standard"), ensure_ascii=False)
    if "https://example.com/path?id=123" not in safe_url_rendered:
        errors.append("standard_export_unsafe_entity_website: safe Entity website missing")

    safe_domain_graph = _base()
    safe_domain_graph["entities"][0]["website"] = None
    safe_domain_graph["entities"][0]["domain"] = "example.com"
    safe_domain_rendered = json.dumps(build_sheets(safe_domain_graph, {"issues": []}, "standard"), ensure_ascii=False)
    if "example.com" not in safe_domain_rendered:
        errors.append("standard_export_unsafe_entity_website: safe Entity domain missing")
    return errors


def _stored_unauthorized_manifest(graph: dict[str, Any]) -> None:
    graph["runs"][0]["review_mode"] = "not_run"
    current_hash = graph_hash(graph)
    graph["audits"].append({
        "audit_id": "audit_001",
        "audited_at": "2026-01-01T00:00:00Z",
        "research_graph_hash": current_hash,
        "audit_graph_hash": current_hash,
        "review_cycle_id": "review_run_001",
        "review_attestation_id": None,
        "reviewed_subject_hash": current_hash,
        "review_provenance_level": "not_run",
        "audit_status": "passed",
        "delivery_status": "standard_development_list",
        "allowed_delivery_statuses": ["initial_lead_list", "standard_development_list"],
        "disclosure_required": False,
        "ok": True,
        "issue_count": 0,
        "issues": [],
    })
    graph["delivery_manifests"].append({
        "delivery_manifest_id": "manifest_001",
        "run_id": "run_001",
        "brief_id": "brief_001",
        "plan_id": "plan_001",
        "audit_id": "audit_001",
        "audit_graph_hash": current_hash,
        "research_graph_hash": current_hash,
        "review_cycle_id": "review_run_001",
        "review_attestation_id": None,
        "reviewed_subject_hash": current_hash,
        "review_provenance_level": "not_run",
        "generated_at": "2026-01-01T00:00:00Z",
        "delivery_status": "standard_development_list",
        "output_mode": "standard",
        "exported_entity_ids": ["ent_001"],
        "exported_contact_ids": ["contact_001"],
        "exported_contact_claim_ids": ["cc_001"],
        "exported_assessment_ids": ["assess_001"],
        "output_files": [],
        "warnings": [],
        "disclosures": [],
    })


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--suite",
        choices=("default", "deep", "all"),
        default="all",
        help="default: discovery export safety; deep: formal-delivery regressions; all: both (default).",
    )
    return parser.parse_args()


def main() -> int:
    suite = parse_args().suite
    if suite == "default":
        errors = _assert_default_discovery_export_filters_unsafe_urls()
        if errors:
            print("Advanced default gate regressions failed:")
            print("\n\n".join(errors))
            return 1
        print("advanced gate regressions passed: suite=default groups=1 failures=0")
        return 0

    tests: list[tuple[str, Callable[[dict[str, Any]], None]]] = []

    def add(name: str, mutate: Callable[[dict[str, Any]], None]) -> None:
        tests.append((name, mutate))

    add("claim_value_not_in_observation", lambda g: g["claims"][0].update({"typed_value": {"text": "unrelated high-risk equipment"}}))
    add("claim_semantic_fields_unanchored", lambda g: g["claims"][0].update({"subject": "Other Buyer", "predicate": "is exclusive purchaser of", "claim_type": "exclusive_purchase_authority", "claim_scope": "global_2026", "typed_value": {"text": "sample product", "exclusive": True}}))
    add("assessment_rationale_unanchored_fact", lambda g: g["assessments"][0].update({"rationale_structured": {"summary": "confirmed exclusive purchaser"}}))
    def claim_value_substring(graph: dict[str, Any]) -> None:
        graph["observations"][0]["raw_excerpt"] = graph["observations"][0]["raw_excerpt"].replace("sample product", "sample products")
        graph["claims"][0]["typed_value"] = {"text": "product"}
        graph["claim_evidence"][0]["claim_field_anchors"]["claim_type"] = "Example Buyer sells sample products"
        graph["claim_evidence"][0]["claim_field_anchors"]["typed_value"] = "product"

    add("claim_value_substring", claim_value_substring)

    def translation_missing_origin(graph: dict[str, Any]) -> None:
        graph["observations"][0].update({"translation_status": "translated", "derived_from_observation_id": "obs_missing"})

    add("translation_missing_origin", translation_missing_origin)

    def translation_cross_entity(graph: dict[str, Any]) -> None:
        graph["entities"].append({"entity_id": "ent_002", "name": "Other Buyer", "website": "https://other.example"})
        origin = copy.deepcopy(graph["observations"][0])
        origin.update({"observation_id": "obs_002", "entity_id": "ent_002"})
        graph["observations"].append(origin)
        graph["observations"][0].update({"translation_status": "translated", "derived_from_observation_id": "obs_002"})

    add("translation_cross_entity_origin", translation_cross_entity)

    def translation_cycle(graph: dict[str, Any]) -> None:
        translated = copy.deepcopy(graph["observations"][0])
        translated.update({"observation_id": "obs_002", "translation_status": "translated", "derived_from_observation_id": "obs_001"})
        graph["observations"].append(translated)
        graph["observations"][0].update({"translation_status": "translated", "derived_from_observation_id": "obs_002"})

    add("translation_cycle_without_original_root", translation_cycle)
    add("unknown_capability_support", lambda g: g["observations"][0].update({"capability": "vendor.magic_lookup"}))
    add("search_result_contact_source", lambda g: g["contact_points"][0].update({"source_type": "search_result"}))

    def contact_form_truncation(graph: dict[str, Any]) -> None:
        graph["observations"][0]["raw_excerpt"] += " Form: https://example.com/contact-form"
        graph["contact_points"][0].update({
            "contact_type": "contact_form",
            "source_literal": "https://example.com/contact-form",
            "normalized_value": "https://example.com/contact",
        })
        graph["contact_claims"][0]["association_evidence_text"] = "Form: https://example.com/contact-form"

    add("contact_form_truncation", contact_form_truncation)

    def ready_without_entity(graph: dict[str, Any]) -> None:
        graph["contact_claims"][0].update({"entity_id": None, "person_id": "person_001"})

    add("ready_contact_without_resolved_entity", ready_without_entity)
    add("manual_contact_fabricated_association", lambda g: g["contact_claims"][0].update({"export_status": "needs_manual_association_review", "user_status": "待确认归属", "association_evidence_text": "Invented association text"}))
    add("exportable_contact_without_entity", lambda g: g["contact_claims"][0].update({"export_status": "export_with_source_note", "user_status": "建议核查后使用", "entity_id": None, "person_id": "person_001"}))

    def email_disguised_as_other(graph: dict[str, Any]) -> None:
        graph["observations"][0]["raw_excerpt"] = "Example Buyer sells sample product. Example Buyer Contact: not-sales@example.com"
        graph["contact_points"][0].update({"contact_type": "other", "source_literal": "sales@example.com", "normalized_value": "sales@example.com"})
        graph["contact_claims"][0]["association_evidence_text"] = "Example Buyer Contact: sales@example.com"

    add("email_disguised_as_other_substring", email_disguised_as_other)

    def cross_company_contact_section(graph: dict[str, Any]) -> None:
        graph["entities"].append({"entity_id": "ent_002", "name": "Other Buyer", "website": "https://other.example"})
        graph["observations"][0]["raw_excerpt"] = "Example Buyer sells sample product. Other Buyer Contact: sales@example.com"
        graph["contact_claims"][0]["association_evidence_text"] = "Other Buyer Contact: sales@example.com"

    add("cross_company_contact_section", cross_company_contact_section)
    add("invalid_claim_support_source_url", lambda g: g["sources"][0].update({"canonical_url": "data:text/html,not-a-source", "final_url": "javascript:alert(1)"}))

    def forged_relationship_evidence(graph: dict[str, Any]) -> None:
        graph["entities"].append({"entity_id": "ent_002", "name": "Other Buyer", "website": "https://other.example"})
        graph["entity_relationships"].append({
            "entity_relationship_id": "rel_001",
            "source_entity_id": "ent_001",
            "target_entity_id": "ent_002",
            "relationship_type": "dealer_of",
            "resolution_status": "contextual",
            "confidence": "high",
            "rationale": "fixture",
            "evidence_claim_ids": ["claim_missing"],
            "evidence_observation_ids": ["obs_missing"],
        })

    add("entity_relationship_forged_evidence", forged_relationship_evidence)
    add("stored_manifest_delivery_status_not_allowed", _stored_unauthorized_manifest)

    def manifest_empty_reference(graph: dict[str, Any]) -> None:
        graph["delivery_manifests"].append({
            "delivery_manifest_id": "manifest_001", "run_id": "", "brief_id": "", "plan_id": "", "audit_id": "",
            "audit_graph_hash": "", "research_graph_hash": "", "review_cycle_id": "review_run_001", "review_attestation_id": None, "reviewed_subject_hash": "a" * 64, "review_provenance_level": "not_applicable",
            "generated_at": "2026-01-01T00:00:00Z", "delivery_status": "standard_development_list", "output_mode": "standard",
            "exported_entity_ids": [], "exported_contact_ids": [], "exported_contact_claim_ids": [], "exported_assessment_ids": [],
            "output_files": [], "warnings": [], "disclosures": [],
        })

    add("manifest_empty_reference_chain", manifest_empty_reference)
    add("unknown_control_field", lambda g: g["runs"][0].update({"audit_bypass": True}))

    errors: list[str] = []
    with tempfile.TemporaryDirectory(prefix="superleads_advanced_gates_") as temporary:
        directory = Path(temporary)
        for name, mutate in tests:
            graph = _base()
            mutate(graph)
            errors.extend(_must_block(name, graph, directory))
        errors.extend(_assert_hold_value_is_not_exported(directory))
        errors.extend(_assert_hold_free_text_is_redacted(directory))
        errors.extend(_assert_manual_contact_is_not_exported(directory))
        errors.extend(_assert_standard_docs_include_source_links())
        errors.extend(_assert_user_file_direct_pressure_tests(directory))
        errors.extend(_assert_material_role_direct_pressure_tests(directory))
        errors.extend(_assert_inquiry_export_redaction(directory))
        errors.extend(_assert_mail_adapter_boundary(directory))
        errors.extend(_assert_platform_and_public_url_pressure_tests(directory))
        errors.extend(_assert_standard_export_filters_unsafe_entity_websites())
        errors.extend(_assert_self_review_disclosure(directory))
        errors.extend(_assert_historical_review_cannot_approve(directory))
        errors.extend(_assert_historical_assessment_cannot_be_reused(directory))
        errors.extend(_assert_formal_exception_bindings(directory))
        errors.extend(_assert_identity_literal_bindings(directory))
        errors.extend(_assert_background_research_contract(directory))
        errors.extend(_assert_background_report_export(directory))
        errors.extend(_assert_phase2_provenance_and_searchlog(directory))
        errors.extend(_assert_target_geography_contract_required(directory))
        errors.extend(_assert_legacy_review_fields_rejected(directory))
        errors.extend(_assert_review_attestation_coverage_and_disclosures(directory))
        invalid_schema = _base()
        invalid_schema["observations"][0]["capability"] = "vendor.magic_lookup"
        schema_path = directory / "schema_fail_closed.json"
        _write_graph(schema_path, invalid_schema)
        fallback = _run([sys.executable, "-S", "-B", str(SCRIPTS / "audit_delivery.py"), str(schema_path), "--delivery-status", "standard_development_list"])
        if fallback.returncode == 0:
            errors.append(f"schema_fail_closed: audit passed without jsonschema\n{fallback.stdout}")

    if errors:
        print("Advanced gate regressions failed:")
        print("\n\n".join(errors))
        return 1
    group_count = len(tests) + 21
    if suite == "all":
        default_errors = _assert_default_discovery_export_filters_unsafe_urls()
        if default_errors:
            print("Advanced default gate regressions failed:")
            print("\n\n".join(default_errors))
            return 1
        group_count += 1
    print(f"advanced gate regressions passed: suite={suite} groups={group_count} failures=0")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
